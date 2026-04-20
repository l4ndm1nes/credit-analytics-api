from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook

from app.application.dto import PlanRowDTO
from app.domain.exceptions import InvalidPlansFileError, ValidationIssue

_REQUIRED_COLUMNS: tuple[str, ...] = ("period", "category", "sum")


class PlansExcelParser:
    def parse(self, content: bytes) -> list[PlanRowDTO]:
        try:
            workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise InvalidPlansFileError(f"cannot open excel file: {exc}") from exc

        sheet = workbook.active
        if sheet is None:
            raise InvalidPlansFileError("workbook has no active sheet")

        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise InvalidPlansFileError("file is empty") from exc

        header = self._normalize_header(header_row)
        missing = [col for col in _REQUIRED_COLUMNS if col not in header]
        if missing:
            raise InvalidPlansFileError(
                f"missing required columns: {', '.join(missing)}",
                issues=[
                    ValidationIssue(location="header", message=f"missing column '{col}'")
                    for col in missing
                ],
            )

        period_idx = header.index("period")
        category_idx = header.index("category")
        sum_idx = header.index("sum")

        parsed: list[PlanRowDTO] = []
        issues: list[ValidationIssue] = []

        for excel_row_number, raw in enumerate(rows_iter, start=2):
            if raw is None or all(cell is None for cell in raw):
                continue

            period_value = self._cell(raw, period_idx)
            category_value = self._cell(raw, category_idx)
            sum_value = self._cell(raw, sum_idx)

            period = self._parse_period(period_value, excel_row_number, issues)
            category = self._parse_category(category_value, excel_row_number, issues)
            amount = self._parse_sum(sum_value, excel_row_number, issues)

            if period is not None and category is not None and amount is not None:
                parsed.append(PlanRowDTO(period=period, category_name=category, sum=amount))

        if issues:
            raise InvalidPlansFileError("invalid plans file", issues=issues)

        if not parsed:
            raise InvalidPlansFileError("file contains no data rows")

        return parsed

    @staticmethod
    def _normalize_header(header_row: tuple[object, ...]) -> list[str]:
        return [str(cell).strip().lower() if cell is not None else "" for cell in header_row]

    @staticmethod
    def _cell(row: tuple[object, ...], index: int) -> object | None:
        if index >= len(row):
            return None
        return row[index]

    @staticmethod
    def _parse_period(
        value: object | None,
        row_number: int,
        issues: list[ValidationIssue],
    ) -> date | None:
        if value is None or (isinstance(value, str) and not value.strip()):
            issues.append(
                ValidationIssue(location=f"row {row_number}, period", message="value is required")
            )
            return None

        period: date | None = None
        if isinstance(value, datetime):
            period = value.date()
        elif isinstance(value, date):
            period = value
        elif isinstance(value, str):
            for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
                try:
                    period = datetime.strptime(value.strip(), fmt).date()
                    break
                except ValueError:
                    continue

        if period is None:
            issues.append(
                ValidationIssue(
                    location=f"row {row_number}, period",
                    message=f"cannot parse date value: {value!r}",
                )
            )
            return None

        if period.day != 1:
            issues.append(
                ValidationIssue(
                    location=f"row {row_number}, period",
                    message="period must be the first day of a month",
                )
            )
            return None

        return period

    @staticmethod
    def _parse_category(
        value: object | None,
        row_number: int,
        issues: list[ValidationIssue],
    ) -> str | None:
        if value is None or (isinstance(value, str) and not value.strip()):
            issues.append(
                ValidationIssue(location=f"row {row_number}, category", message="value is required")
            )
            return None
        return str(value).strip()

    @staticmethod
    def _parse_sum(
        value: object | None,
        row_number: int,
        issues: list[ValidationIssue],
    ) -> Decimal | None:
        if value is None or (isinstance(value, str) and not value.strip()):
            issues.append(
                ValidationIssue(location=f"row {row_number}, sum", message="value is required")
            )
            return None

        try:
            amount = Decimal(str(value).replace(",", "."))
        except (InvalidOperation, ValueError):
            issues.append(
                ValidationIssue(
                    location=f"row {row_number}, sum",
                    message=f"cannot parse numeric value: {value!r}",
                )
            )
            return None

        if amount < 0:
            issues.append(
                ValidationIssue(
                    location=f"row {row_number}, sum",
                    message="sum must be non-negative",
                )
            )
            return None

        return amount
